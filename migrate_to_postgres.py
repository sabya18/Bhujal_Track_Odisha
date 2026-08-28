import openpyxl
import os
import psycopg2
from psycopg2.extensions import ISOLATION_LEVEL_AUTOCOMMIT
import re
import datetime

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
FILE_PATH = os.path.join(SCRIPT_DIR, "CTC Pre-monsoon Field Book 2026.xlsx")

def dms_to_dd(dms_str):
    if dms_str is None:
        return None
    dms_str = str(dms_str).strip()
    if not dms_str:
        return None
    
    if any(delim in dms_str for delim in ['_', '-', ' ', '°', '\'', '"', ':']):
        cleaned = dms_str.replace('_', ' ').replace('-', ' ').replace('°', ' ').replace('\'', ' ').replace('"', ' ').replace(':', ' ')
        parts = [p.strip() for p in cleaned.split() if p.strip()]
        if len(parts) >= 1:
            try:
                d = float(parts[0])
                m = float(parts[1]) if len(parts) > 1 else 0.0
                s = float(parts[2]) if len(parts) > 2 else 0.0
                
                direction = 1
                for p in parts:
                    if p.upper() in ['S', 'W']:
                        direction = -1
                
                dd = d + m / 60.0 + s / 3600.0
                return round(dd * direction, 6)
            except Exception:
                pass
                
    try:
        return float(dms_str)
    except ValueError:
        pass
        
    return None

def format_date(val):
    if val is None:
        return None
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime('%d.%m.%Y')
    val_str = str(val).strip()
    if not val_str or val_str.lower() == 'nan' or val_str.lower() == 'none':
        return None
    if ' ' in val_str:
        val_str = val_str.split(' ')[0]
    if '-' in val_str:
        parts = val_str.split('-')
        if len(parts) == 3:
            if len(parts[0]) == 4:
                return f"{parts[2]}.{parts[1]}.{parts[0]}"
            return f"{parts[0]}.{parts[1]}.{parts[2]}"
    return val_str

def get_column_mapping(ws):
    header_row = None
    for r in range(1, 15):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        vals_str = [str(v).lower() for v in vals if v is not None]
        has_well = any("well number" in v or "well no" in v or "well id" in v or "well_id" in v for v in vals_str)
        has_loc = any("location of" in v or "location" in v for v in vals_str)
        if has_well and has_loc:
            header_row = r
            break
            
    if not header_row:
        for r in range(1, 10):
            vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            vals_str = [str(v).lower() for v in vals if v is not None]
            if any("location" in v for v in vals_str) and any("block" in v or "sl" in v for v in vals_str):
                header_row = r
                break
                
    if not header_row:
        return None, {}
        
    headers = [str(ws.cell(row=header_row, column=c).value).strip() if ws.cell(row=header_row, column=c).value is not None else "" for c in range(1, ws.max_column + 1)]
    
    col_map = {}
    for idx, h in enumerate(headers):
        h_lower = h.lower()
        if "sl" in h_lower and "no" in h_lower:
            col_map["sl_no"] = idx
        elif "block" in h_lower or "urban area" in h_lower or "urban_area" in h_lower:
            col_map["block"] = idx
        elif "location" in h_lower:
            col_map["location"] = idx
        elif "well type" in h_lower:
            col_map["well_type"] = idx
        elif "well number" in h_lower or "well no" in h_lower or "well id" in h_lower or "well_id" in h_lower:
            col_map["well_number"] = idx
        elif "lat" in h_lower and "date" not in h_lower:
            col_map["lat"] = idx
        elif ("long" in h_lower or "lon" in h_lower) and "date" not in h_lower:
            col_map["lon"] = idx
        elif "dt_site" in h_lower or "dt_sitevisit" in h_lower or "date of site" in h_lower or "dt_sitevist" in h_lower:
            col_map["date"] = idx
        elif "total depth" in h_lower or "tot_ depth" in h_lower or "total_depth" in h_lower or "depth" in h_lower:
            if "parapet" not in h_lower:
                col_map["depth"] = idx
        elif "parapet" in h_lower:
            col_map["parapet"] = idx
        elif "dtgwl" in h_lower and "bmp" in h_lower:
            col_map["dtgwl_bmp"] = idx
        elif "dtgwl" in h_lower and "mbgl" in h_lower:
            col_map["dtgwl_mbgl"] = idx
        elif "remark" in h_lower:
            col_map["remarks"] = idx
            
    if "sl_no" not in col_map: col_map["sl_no"] = 0
    if "block" not in col_map: col_map["block"] = 1
    if "location" not in col_map: col_map["location"] = 2
    if "well_type" not in col_map: col_map["well_type"] = 3
    if "well_number" not in col_map: col_map["well_number"] = 4
    if "lat" not in col_map: col_map["lat"] = 5
    if "lon" not in col_map: col_map["lon"] = 6
    if "date" not in col_map: col_map["date"] = 7
    if "depth" not in col_map: col_map["depth"] = 8
    if "parapet" not in col_map: col_map["parapet"] = 9
    if "dtgwl_bmp" not in col_map: col_map["dtgwl_bmp"] = 10
    if "remarks" not in col_map: col_map["remarks"] = 13
    
    mbgl_col = None
    for idx, h in enumerate(headers):
        if "mbgl" in h.lower():
            mbgl_col = idx
    if mbgl_col is not None:
        col_map["dtgwl_mbgl"] = mbgl_col
        
    return header_row, col_map

def main():
    # 1. Connect to PostgreSQL default DB and create the database
    conn = psycopg2.connect(
        host="localhost",
        user="postgres",
        password="1234",
        dbname="postgres"
    )
    conn.set_isolation_level(ISOLATION_LEVEL_AUTOCOMMIT)
    cursor = conn.cursor()
    
    # Check database existence
    cursor.execute("SELECT 1 FROM pg_database WHERE datname = 'bhujal_monitor';")
    exists = cursor.fetchone()
    if not exists:
        print("Creating database 'bhujal_monitor'...")
        cursor.execute("CREATE DATABASE bhujal_monitor;")
    else:
        print("Database 'bhujal_monitor' already exists.")
        
    cursor.close()
    conn.close()
    
    # 2. Connect to the target database and build schema
    conn = psycopg2.connect(
        host="localhost",
        user="postgres",
        password="1234",
        dbname="bhujal_monitor"
    )
    conn.set_isolation_level(ISOLATION_LEVEL_AUTOCOMMIT)
    cursor = conn.cursor()
    
    # Drops existing tables to start fresh
    cursor.execute("DROP TABLE IF EXISTS visits CASCADE;")
    cursor.execute("DROP TABLE IF EXISTS wells CASCADE;")
    cursor.execute("DROP TABLE IF EXISTS app_users CASCADE;")
    
    # Construct app_users table
    cursor.execute("""
        CREATE TABLE app_users (
            username VARCHAR(50) PRIMARY KEY,
            password VARCHAR(255) NOT NULL
        );
    """)
    
    # Insert default users
    cursor.execute("""
        INSERT INTO app_users (username, password)
        VALUES ('gwd_officer', 'gwd_password_2026')
        ON CONFLICT (username) DO NOTHING;
    """)
    
    # Construct wells table
    cursor.execute("""
        CREATE TABLE wells (
            well_number VARCHAR(50) PRIMARY KEY,
            sheet VARCHAR(100),
            district VARCHAR(100),
            block VARCHAR(150),
            location TEXT,
            well_type VARCHAR(20),
            lat_raw VARCHAR(50),
            lon_raw VARCHAR(50),
            lat DOUBLE PRECISION,
            lon DOUBLE PRECISION,
            depth VARCHAR(50),
            parapet_height DOUBLE PRECISION DEFAULT 0.0,
            msl DOUBLE PRECISION,
            rl DOUBLE PRECISION
        );
    """)
    
    # Construct visits table
    cursor.execute("""
        CREATE TABLE visits (
            id SERIAL PRIMARY KEY,
            well_number VARCHAR(50) REFERENCES wells(well_number) ON DELETE CASCADE,
            season_key VARCHAR(50),
            date VARCHAR(20),
            dtgwl_bmp DOUBLE PRECISION,
            dtgwl_mbgl DOUBLE PRECISION,
            remarks TEXT,
            UNIQUE(well_number, season_key)
        );
    """)
    
    # Indexes for spatial optimization and queries
    cursor.execute("CREATE INDEX idx_wells_block ON wells(block);")
    cursor.execute("CREATE INDEX idx_visits_lookup ON visits(well_number, season_key);")
    
    print("Schema initialized and default user seeded successfully.")
    
    # 3. Read the Excel data and insert into tables
    if not os.path.exists(FILE_PATH):
        print(f"Error: Excel workbook not found at {FILE_PATH}")
        return
        
    wb = openpyxl.load_workbook(FILE_PATH, data_only=True)
    
    wells_inserted = 0
    visits_inserted = 0
    
    for sheet_name in wb.sheetnames:
        # Ignore non-district data sheets
        if sheet_name in ['Rules', 'Index', '1']:
            continue
            
        ws = wb[sheet_name]
        header_row, col_map = get_column_mapping(ws)
        if not header_row:
            continue
            
        print(f"Importing sheet: {sheet_name}...")
        
        # Determine district from sheet name
        district = "Cuttack"
        s = sheet_name.lower().strip()
        if 'kendrapara' in s:
            district = 'Kendrapara'
        elif 'jajpur' in s:
            district = 'Jajpur'
        elif 'jspur' in s or 'jagatsinghpur' in s:
            district = 'Jagatsinghpur'
            
        for r in range(header_row + 1, ws.max_row + 1):
            c1 = ws.cell(row=r, column=1).value
            c2 = ws.cell(row=r, column=2).value
            if c1 == 1 and c2 == 2:
                continue
                
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            if all(v is None for v in row_vals):
                continue
                
            def val(key):
                col_idx = col_map.get(key)
                if col_idx is not None and col_idx < len(row_vals):
                    return row_vals[col_idx]
                return None
                
            w_num = val("well_number")
            if not w_num:
                continue
            w_num_str = str(w_num).strip().upper()
            if not re.search(r'[A-Za-z]', w_num_str):
                continue
                
            lat_raw = val("lat")
            lon_raw = val("lon")
            lat_dd = dms_to_dd(lat_raw)
            lon_dd = dms_to_dd(lon_raw)
            
            p_height = val("parapet")
            try:
                p_height = float(p_height) if p_height is not None else 0.0
            except ValueError:
                p_height = 0.0
                
            bmp_val = val("dtgwl_bmp")
            try:
                bmp_val = float(bmp_val) if bmp_val is not None else None
            except ValueError:
                bmp_val = None
                
            mbgl_val = val("dtgwl_mbgl") if "dtgwl_mbgl" in col_map else None
            try:
                mbgl_val = float(mbgl_val) if mbgl_val is not None else None
            except ValueError:
                mbgl_val = None
                
            if mbgl_val is None and bmp_val is not None:
                mbgl_val = round(bmp_val - p_height, 2)
                
            rem = val("remarks")
            rem_str = str(rem).strip() if rem is not None else ""
            
            # Use Pre-monsoon 2026 as default season key for this import
            season_key = "2026_PreMon"
            date_formatted = format_date(val("date"))
            
            # Fetch MSL and calculate RL if missing
            msl = None
            rl = None
            
            # Insert station metadata (ignore duplicates)
            try:
                cursor.execute("""
                    INSERT INTO wells (
                        well_number, sheet, district, block, location, well_type,
                        lat_raw, lon_raw, lat, lon, depth, parapet_height, msl, rl
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (well_number) DO NOTHING;
                """, (
                    w_num_str, sheet_name, district, 
                    str(val("block")).strip() if val("block") else "",
                    str(val("location")).strip() if val("location") else "",
                    str(val("well_type")).strip() if val("well_type") else "",
                    str(lat_raw).strip() if lat_raw else "",
                    str(lon_raw).strip() if lon_raw else "",
                    lat_dd, lon_dd,
                    str(val("depth")).strip() if val("depth") else "",
                    p_height, msl, rl
                ))
                wells_inserted += 1
            except Exception as e:
                print(f"Error inserting well {w_num_str}: {e}")
                
            # Insert visit record if measurement is valid
            if date_formatted and (bmp_val is not None or mbgl_val is not None):
                try:
                    cursor.execute("""
                        INSERT INTO visits (well_number, season_key, date, dtgwl_bmp, dtgwl_mbgl, remarks)
                        VALUES (%s, %s, %s, %s, %s, %s)
                        ON CONFLICT (well_number, season_key) DO UPDATE
                        SET date = EXCLUDED.date,
                            dtgwl_bmp = EXCLUDED.dtgwl_bmp,
                            dtgwl_mbgl = EXCLUDED.dtgwl_mbgl,
                            remarks = EXCLUDED.remarks;
                    """, (w_num_str, season_key, date_formatted, bmp_val, mbgl_val, rem_str))
                    visits_inserted += 1
                except Exception as e:
                    print(f"Error inserting visit for {w_num_str}: {e}")
                    
    cursor.close()
    conn.close()
    print(f"\nMigration complete! Synced {wells_inserted} stations and {visits_inserted} visits to PostgreSQL database.")

if __name__ == "__main__":
    main()
